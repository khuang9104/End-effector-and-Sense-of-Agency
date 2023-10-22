"""
Abstract control experiment.

During calibration phase, the participant is asked to reproduce a series of
muscle contractions to determine noise level and maximum contraction during the
control. Data are collected and stored so that they can be used later to 
normalize control values.

During the real-time control experiments, participants are controlling a cursor 
through their muscle activity, and are asked to bring that cursor into the 
targets presented on the screen. Raw data, processed data, and outcome measured 
are saved for later analysis.

Input devices:
trigno
    Delsys Trigno EMG system.
myo
    Myo armband.
noise
    Noise generator
quattro
    Delsys quattro EMG system


All configuration settings are stored and loaded from an external configuration
file (``config.ini``).
"""

import os
import numpy as np
import h5py
import openpyxl
import socket
import threading
from time import localtime, strftime, sleep
from datetime import datetime

from argparse import ArgumentParser
from configparser import ConfigParser
from scipy.signal import butter

from axopy.experiment import Experiment
from axopy.task import Task
from axopy import util
from axopy.gui.graph import SignalWidget
from axopy.pipeline import (Windower, Pipeline, Filter,
                            FeatureExtractor, Ensure2D, Block, Callable)
from graphics import CalibWidget 
from axopy.features import MeanAbsoluteValue
from axopy.daq import Keyboard


class Normalize(Block):
    def process(self, data):
        data_norm = (data - c_min)/(c_max-c_min)
        data_norm[data_norm < 0] = 0.0
        return data_norm

class _BaseTask(Task):
    """Base experimental task.

    Implements the processing pipeline, the daqstream and the trial counter.
    """
    def __init__(self):
        super(_BaseTask, self).__init__()
        self.pipeline = self.make_pipeline()

    def make_pipeline(self):
        b,a = butter(FILTER_ORDER,
                      (LOWCUT / S_RATE / 2., HIGHCUT / S_RATE / 2.),
                      'bandpass')
        pipeline = Pipeline([
            Windower(int(S_RATE * WIN_SIZE)),
            Filter(b, a=a,
                   overlap=(int(S_RATE * WIN_SIZE) -
                            int(S_RATE * READ_LENGTH))),
            FeatureExtractor([('MAV', MeanAbsoluteValue())],
                             n_channels),
            Ensure2D(orientation='row'),
            Normalize()
        ])

        print('Pipeline ok...')
        return pipeline

    def prepare_daq(self, daqstream):
        self.daqstream = daqstream
        self.daqstream.start()

    def key_press(self, key):
        super(_BaseTask, self).key_press(key)
        if key == util.key_escape:
            self.finish()

    def finish(self):
        self.daqstream.stop()
        self.finished.emit()

class DataCollection(_BaseTask):
    """Data collection task for callibration of the system.

    Collects minimum and maximum values for interface. Raw EMG activity is
    shown on screeen. Experimenter decides through key presses when muscles
    at rest and when they represent full activity.
    """

    def __init__(self):
        super(DataCollection, self).__init__()
        (self.pipeline_raw, self.pipeline_calib) = self.make_pipeline_calib()

    def make_pipeline_calib(self):
        print('Pipeline calib ok...')
        b,a = butter(FILTER_ORDER,
                      (LOWCUT / S_RATE / 2., HIGHCUT / S_RATE / 2.),
                      'bandpass')
        pipeline_raw = Pipeline([
            Windower(int(S_RATE * WIN_SIZE_CALIB)),
            Filter(b, a=a,
                   overlap=(int(S_RATE * WIN_SIZE_CALIB) -
                            int(S_RATE * READ_LENGTH))),
            Ensure2D(orientation='col')
        ])

        pipeline_calib = Pipeline([
            Windower(int(S_RATE * WIN_SIZE)),
            Filter(b, a=a,
                   overlap=(int(S_RATE * WIN_SIZE) -
                            int(S_RATE * READ_LENGTH))),
            FeatureExtractor([('MAV', MeanAbsoluteValue())],
                    n_channels),
            Ensure2D(orientation='col'),
            Windower(int((1/READ_LENGTH) * WIN_SIZE_CALIB)),
        ])

        return pipeline_raw, pipeline_calib


    def prepare_design(self, design):
        for b in range(N_BLOCKS):
            block = design.add_block()
            for t in range(N_TRIALS):
                block.add_trial()

    def prepare_graphics(self, container):
        self.scope = CalibWidget(channel_names, c_min, c_max, c_std, c_select, channels_task)

    def prepare_storage(self, storage):
        time = strftime('%Y%m%d%H%M%S', localtime())
        self.writer = storage.create_task('calibration' + '_' + time)
        self.data_dir = os.path.join(os.path.dirname(
                            os.path.realpath(__file__)),
                            'data')
        
        # Save Config File
        block_savedir = os.path.join(self.data_dir, exp.subject, 
                    'calibration' + '_' + time)
        config = ConfigParser()
        config.read(CONFIG)
        with open(block_savedir+"\\config.ini", 'w') as f:
            config.write(f)

    def run_trial(self, trial):
        # saving data
        trial.add_array('data_raw', stack_axis=1)
        trial.add_array('data_proc', stack_axis=1)
        trial.add_array('c_min', stack_axis=1)
        trial.add_array('c_max', stack_axis=1)
        trial.add_array('c_std', stack_axis=1)
        trial.add_array('c_select', stack_axis=1)

        self.pipeline.clear()
        self.connect(self.daqstream.updated, self.update)

    def update(self, data):
        if self.pipeline_raw is not None:
            data_raw = self.pipeline_raw.process(data)
        if self.pipeline is not None:
            data_proc = self.pipeline.process(data)
        if self.pipeline_calib is not None:
            data_calib = self.pipeline_calib.process(data)

        self.scope.plot(data_raw, data_proc, data_calib)
        # Update Arrays
        self.trial.arrays['data_raw'].stack(data)
        self.trial.arrays['data_proc'].stack(data_proc)

    def key_press(self, key):
        super(_BaseTask, self).key_press(key)
        if key == util.key_escape:
            self.finish()

    def finish(self):
        self.trial.arrays['c_min'].stack(c_min)
        self.trial.arrays['c_max'].stack(c_max)
        self.trial.arrays['c_std'].stack(c_std)
        self.trial.arrays['c_select'].stack(c_select)
        self.writer.write(self.trial)
        self.disconnect(self.daqstream.updated, self.update)
        self.daqstream.stop()
        self.finished.emit()

class RealTimeControl(_BaseTask):
    """Real time abstract control
    """
    def __init__(self):
        super(RealTimeControl, self).__init__()
        self.make_display_pipeline()

    def prepare_design(self, design):
        for b in range(1):
            block = design.add_block()
            for t in range(1):
                block.add_trial()

    def make_display_pipeline(self):
        # windower to display something meaningful in oscilloscope
        self.display_pipeline = Pipeline([
            Windower(500)
        ])
    
    def prepare_graphics(self, container):
        self.scope = SignalWidget(channel_names=channel_names, yrange=(0,1))
        container.set_widget(self.scope)

    def prepare_storage(self, storage):
        time = strftime('%Y%m%d%H%M%S', localtime())
        self.writer = storage.create_task('control' + '_' + time)
        self.data_dir = os.path.join(os.path.dirname(
                            os.path.realpath(__file__)),
                            'data')
        
        # Save Config File
        block_savedir = os.path.join(self.data_dir, exp.subject, 
                            'control' + '_' + time)
        config = ConfigParser()
        config.read(CONFIG)
        with open(block_savedir+"\\config.ini", 'w') as f:
            config.write(f)

    def run_trial(self, trial):
        trial.add_array('data_raw', stack_axis=1)
        trial.add_array('data_proc', stack_axis=1)

        self.display_pipeline.clear()
        self.connect(self.daqstream.updated, self.update_trial)

    def update_trial(self, data):
        # for EMGsim test only -Kang
        #if args.emgsim and data > 0:
        #    print('data > 0: ' + datetime.utcnow().isoformat(sep=' ', timespec='milliseconds'))
            
        # process data
        data_proc = self.pipeline.process(data)
        
        # update processed_data for rest check -Kang
        global processed_data
        processed_data = data_proc

        # update scope
        data_scope = self.display_pipeline.process(data_proc)
        self.scope.plot(data_scope)
        
        # Threshold visual aid -Kang
        global time_record, active_time, reset_time
        if data_proc > THRESHOLD:
            self.scope.plot_data_items[0].setPen('g')
        else:
            self.scope.plot_data_items[0].setPen('r')
        
        # save data
        #self.trial.arrays['data_raw'].stack(data)
        #self.trial.arrays['data_proc'].stack(np.transpose(data_proc))
        
        # for EMGsim test only -Kang
        #if args.emgsim and data_proc > 0:
        #    print('data_proc > 0: ' + datetime.utcnow().isoformat(sep=' ', timespec='milliseconds'))
  
        # Send active signal to MyoSocket(Java) -Kang
        global sync_flag
        if sync_flag == True and data_proc > THRESHOLD:
            sync_flag = False
            ready_to_send_time = datetime.utcnow().isoformat(sep=' ', timespec='milliseconds')
            print('active signal ready to send: ' + ready_to_send_time)
            signal_ready_times.append(ready_to_send_time.split(":")[2])
            client_data, addr = server_data.accept()
            output = bytes('active', 'utf-8')
            client_data.sendall(output)
            client_data.close()
    
    def finish_trial(self):
        self.writer.write(self.trial)
        # self.disconnect(self.daqstream.updated, self.update)
        self.disconnect(self.daqstream.updated, self.update_trial)

    def finish(self):
        self.writer.write(self.trial)
        # self.disconnect(self.daqstream.updated, self.update)
        self.disconnect(self.daqstream.updated, self.update_trial)
        self.daqstream.stop()
        self.finished.emit()
    
    def key_press(self, key):
        if key == util.key_escape:
            # Save system delay test data (ss.SSS)  -Kang
            if len(signal_ready_times) > 0:
                time = strftime('%Y%m%d%H%M%S', localtime())
                global DATA_PATH
                write_excel(os.path.join(DATA_PATH, "py" + time + ".xlsx"), signal_ready_times)
            self.finish()
        else:
            super().key_press(key)

if __name__ == '__main__':
    parser = ArgumentParser()
    task = parser.add_mutually_exclusive_group(required=True)
    task.add_argument('--train', action='store_true')
    task.add_argument('--test', action='store_true')
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument('--trigno', action='store_true')
    source.add_argument('--myo', action='store_true')
    source.add_argument('--noise', action='store_true')
    source.add_argument('--quattro', action='store_true')
    source.add_argument('--emgsim', action='store_true')
    args = parser.parse_args()

    CONFIG = 'config.ini'
    cp = ConfigParser()
    cp.read(os.path.join(os.path.dirname(os.path.realpath(__file__)),
                         'config.ini'))
    
    SUBJECT = cp.get('participant', 'subject')
    READ_LENGTH = cp.getfloat('hardware', 'read_length')
    CHANNELS = list(map(int, (cp.get('hardware', 'channels').split(','))))
    n_channels = len(CHANNELS)
    WIN_SIZE = cp.getfloat('processing', 'win_size')
    LOWCUT = cp.getfloat('processing', 'lowcut')
    HIGHCUT = cp.getfloat('processing', 'highcut')
    FILTER_ORDER = cp.getfloat('processing', 'filter_order')
    channels_task = cp.getint('hardware', 'channels')
    THRESHOLD = cp.getfloat('setting', 'threshold')
    DATA_PATH = cp.get('data_save_addresss', 'addresss')

    # prepare oscilloscope
    channel_names = ['EMG']
    channel_colours = ['b']

    if args.trigno:
        from pytrigno import TrignoEMG
        S_RATE = 2000.
        dev = TrignoEMG(channels=CHANNELS, zero_based=False,
                        samples_per_read=int(S_RATE * READ_LENGTH))
        
    elif args.noise:
        from axopy.daq import NoiseGenerator
        S_RATE = 2000.
        dev = NoiseGenerator(rate=S_RATE, num_channels=n_channels, amplitude=10.0, read_size=int(S_RATE * READ_LENGTH))

    elif args.quattro:
        from pytrigno import QuattroEMG
        S_RATE = 2000.
        dev = QuattroEMG(sensors=range(1,3), 
                        samples_per_read=int(S_RATE * READ_LENGTH),
                        zero_based=False,
                        mode=313,
                        units='normalized',
                        data_port=50043)

    elif args.emgsim:
        keys = list('a')
        S_RATE = 2000
        dev = Keyboard(rate=S_RATE, keys=keys)
    
    exp = Experiment(daq=dev, subject=SUBJECT, allow_overwrite=False)

    if args.train:
        N_TRIALS = cp.getint('calibration', 'n_trials')
        N_BLOCKS = cp.getint('calibration', 'n_blocks')
        WIN_SIZE_CALIB = cp.getfloat('calibration', 'win_size')
        channel_names = ['EMG ' + str(i) for i in range(1, n_channels+1)]
        c_min = np.zeros(n_channels)
        c_max = np.ones(n_channels)
        c_std = np.zeros(n_channels)
        c_select = np.full(n_channels, np.nan)

        # download existing calibration info
        use_previous_calib = cp.getint('calibration', 'use_previous_calib')
        root_subject = os.path.join(os.path.dirname(os.path.realpath(__file__)),'data', exp.subject)
        if os.path.exists(root_subject) & use_previous_calib:
            subfolders = [f.path for f in os.scandir(root_subject) 
                                if f.is_dir()]
            if subfolders:
                calib_ind = np.array([])
                for i in range(len(subfolders)):
                    # check if existing calib & if there's data
                    if 'calibration_20' in subfolders[i]:
                        if len(os.listdir(subfolders[i])):
                            calib_ind = np.append(calib_ind, i)
                print(calib_ind)
                if calib_ind.size != 0:
                    last_calib = subfolders[int(calib_ind[-1:])]

                    f = h5py.File([last_calib + '\\c_min.hdf5'][0], 'r')
                    c_min = f['0'][0:]
                    print('c_min: ', c_min)
                    f = h5py.File([last_calib + '\\c_max.hdf5'][0], 'r')
                    c_max = f['0'][0:]
                    print('c_max: ', c_max)
                    f = h5py.File([last_calib + '\\c_select.hdf5'][0], 'r')
                    c_select = f['0'][0:]
                    print('previously existing calibration loaded')

        print('Running calibration...')
        exp.run(DataCollection())
    elif args.test:
        # System delay test -Kang
        signal_ready_times = []
        
        # Synchronization flag -Kang
        sync_flag = False
        
        # Current processed data -Kang
        processed_data = 0
        
        # Output socket -Kang
        address_data = ('localhost', 8888)   
        server_data = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server_data.bind(address_data)
        server_data.listen(1) # Max 1 connections
        
        # Socket monitor -Kang
        address = ('localhost', 9999)   
        server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server.bind(address)
        server.listen(1) # Max 1 connections
        
        # Socket Monitor (thread) -Kang
        def monitor():
            while True:
                client, addr = server.accept()
                input_port = client.recv(32)    # 32 byte
                if str(input_port, 'utf-8') == 'Sync':
                    output = bytes('Synced', 'utf-8')
                    client.sendall(output)
                    global sync_flag 
                    sync_flag = True
                elif str(input_port, 'utf-8') == 'Connect':
                    output = bytes('Connected', 'utf-8')
                    client.sendall(output)
                elif str(input_port, 'utf-8') == 'RestCheck':
                    if processed_data > THRESHOLD:
                        output = bytes('NotReady', 'utf-8')
                    else:
                        output = bytes('Ready', 'utf-8')
                    client.sendall(output)    
                client.close()
                
        thread_mon = threading.Thread(target=monitor,args=())
        thread_mon.daemon = True
        thread_mon.start()
        
        # Save system delay test data -Kang
        def write_excel(path, values):
            index = len(values)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "PythonSocket"
            sheet.cell(row=1, column=1, value="Action signal send time (ss.SSS)")
            for i in range(0, index):
                sheet.cell(row=i+2, column=1, value=str(values[i]))
            workbook.save(path)
        
        # download calibration info
        root_subject = os.path.join(os.path.dirname(os.path.realpath(__file__)),'data', exp.subject)
        subfolders = [f.path for f in os.scandir(root_subject) if f.is_dir()]
        calib_ind = np.array([])
        for i in range(len(subfolders)):
            if 'calibration_20' in subfolders[i]:
                calib_ind = np.append(calib_ind, i)
        last_calib = subfolders[int(calib_ind[-1:])]

        f = h5py.File([last_calib + '\\c_min.hdf5'][0], 'r')
        c_min = f['0'][0:]
        print('c_min: ', c_min)
        f = h5py.File([last_calib + '\\c_max.hdf5'][0], 'r')
        c_max = f['0'][0:]
        print('c_max: ', c_max)
        f = h5py.File([last_calib + '\\c_select.hdf5'][0], 'r')
        c_select = f['0'][0:]
        print('c_select: ',c_select)

        # determine control channels
        CONTROL_CHANNELS = np.array([])
        for i in range(channels_task):
            CONTROL_CHANNELS = np.append(CONTROL_CHANNELS, np.where(c_select == i)[0][0])
        CONTROL_CHANNELS = CONTROL_CHANNELS.astype(int)

        # Check whether data save address exist or not -Kang
        if os.path.isdir(DATA_PATH):
            print('Data save path check passes.')
            exp.run(RealTimeControl())
        else:
            print('Data save path is not exist!!!')
            sleep(10)


